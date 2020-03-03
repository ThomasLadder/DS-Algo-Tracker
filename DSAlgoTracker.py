# This script is a simple application designed to help me keep track of Data Structures and Algorithms in preparation
# for job interviews.

import openpyxl

def askQuestions():
    question = input("What was the question?")
    category = input("What was the category?")
    platform = input("What was the platform?")
    notes = input("Any notes?")
    status = input("Did you finish?")

    return question, category, platform, notes, status

#def inputAnswers(question, category, platform, notes, status):


wb = openpyxl.load_workbook('DS+Algo.xlsx')
sheet = wb['Sheet1']
length = 20
count = 0
for i in range(1, length):
    if sheet.cell(row=i, column=1).value is not None:
        continue
    else:
        new_row = i
        break

if new_row:
    question, category, platform, notes, status = askQuestions()
    sheet.cell(row=new_row, column=1).value = question
    sheet.cell(row=new_row, column=2).value = category
    sheet.cell(row=new_row, column=3).value = platform
    sheet.cell(row=new_row, column=4).value = notes
    sheet.cell(row=new_row, column=5).value = status
    wb.save("DS+Algo.xlsx")

else:
    print("no more space!")









